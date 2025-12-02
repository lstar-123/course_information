# parse_course_this_week.py
# -*- coding: utf-8 -*-
# 自动登录教务系统并导出当前周课程表（支持OCR验证码识别）
# 环境依赖: pip install requests beautifulsoup4 pillow pytesseract lxml openpyxl

import os
import time
import urllib.parse
from pathlib import Path
from datetime import datetime, timedelta, timezone

import openpyxl
import requests
from bs4 import BeautifulSoup
from PIL import Image, ImageFilter, ImageEnhance, ImageOps
import pytesseract
import webbrowser
import xlrd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# ---------------- CONFIG ----------------
BASE = "https://jwyth.hnkjxy.net.cn"
LOGIN_PAGE = BASE + "/"
SESS_URL = BASE + "/Logon.do?method=logon&flag=sess"
LOGIN_POST = BASE + "/Logon.do?method=logon"
CAPTCHA_URL = BASE + "/verifycode.servlet"
COURSE_EXPORT_URL = BASE + "/jsxsd/xskb/xskb_print.do"

COMMON_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/142.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Language": "zh-CN,zh;q=0.9",
    "Connection": "keep-alive",
}

USERNAME = os.environ.get("JW_USERNAME")
PASSWORD = os.environ.get("JW_PASSWORD") or ""
if USERNAME is None:
    raise SystemExit("❌ 请先通过环境变量 JW_USERNAME / JW_PASSWORD 提供登录凭证")

# ---------------- UTIL ----------------
def extract_hidden_fields(html):
    soup = BeautifulSoup(html, "html.parser")
    return {i.get("name"): i.get("value", "") for i in soup.select("input[type=hidden]") if i.get("name")}

def make_encoded(username, password, scode, sxh):
    code = f"{username}%%%{password}"
    encoded = ""
    i = 0
    for ch in code:
        if i < len(sxh) and sxh[i].isdigit():
            n = int(sxh[i])
        else:
            n = 0
        encoded += ch + scode[:n]
        scode = scode[n:]
        i += 1
    return encoded

def save_text(path: Path, text: str):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")
    return str(path)

# ---------- OCR 部分 ----------
def preprocess_image(image_path: Path):
    """对验证码图片进行预处理，提高 OCR 识别率"""
    img = Image.open(image_path).convert("L")  # 灰度化
    img = ImageOps.invert(img)  # 反色，白底黑字
    img = img.filter(ImageFilter.MedianFilter())  # 中值滤波去噪
    threshold = 150
    img = img.point(lambda x: 255 if x > threshold else 0)  # 二值化
    return img

def recognize_captcha(image_path: str) -> str:
    """对验证码图片进行预处理并使用OCR识别"""
    try:
        img = Image.open(image_path)

        # 转为灰度图
        img = img.convert("L")

        # 二值化（去背景）
        threshold = 140
        img = img.point(lambda x: 255 if x > threshold else 0)

        # 去除边缘噪点
        img = ImageOps.expand(img, border=5, fill="white")
        img = img.filter(ImageFilter.MedianFilter(size=3))

        # OCR识别
        config = "--psm 7 -c tessedit_char_whitelist=0123456789abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ"
        text = pytesseract.image_to_string(img, config=config)

        # 清洗输出结果
        text = "".join(ch for ch in text.strip() if ch.isalnum())
        if len(text) < 4:  # 验证码一般为4位
            raise ValueError("识别结果过短")
        print(f"🤖 OCR 识别验证码: {text}")
        return text
    except Exception as e:
        print(f"🤖 OCR 识别验证码: [识别失败]（{e}）")
        return ""

def download_captcha_and_ocr(session):
    """下载验证码 -> OCR 识别"""
    r = session.get(CAPTCHA_URL + "?t=" + str(int(time.time())), headers=COMMON_HEADERS, timeout=15)
    r.raise_for_status()
    save_dir = Path(__file__).parent / "captcha_image_library"
    save_dir.mkdir(exist_ok=True)
    filename = f"captcha_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
    save_path = save_dir / filename

    with open(save_path, "wb") as f:
        f.write(r.content)
        f.flush()
    print("🖼 验证码已保存到:", save_path)

    captcha_text = recognize_captcha(save_path)
    if not captcha_text or len(captcha_text) < 4:
        print("⚠️ OCR 识别不稳定，请人工输入:")
        try:
            webbrowser.open("file://" + str(save_path))
        except Exception:
            pass
        captcha_text = input("请输入验证码（区分大小写）：").strip()
    return captcha_text

# ---------------- LOGIN ----------------
def login_via_raw_body():
    s = requests.Session()
    s.headers.update(COMMON_HEADERS)
    print("Step1: GET 登录页")
    r1 = s.get(LOGIN_PAGE, timeout=15)
    save_text(Path("debug") / "debug_loginpage.html", r1.text)

    print("Step2: 获取 scode/sxh")
    r_sess = s.post(SESS_URL, headers=COMMON_HEADERS, timeout=15)
    if "#" not in r_sess.text:
        raise RuntimeError("flag=sess 未返回 scode/sxh")
    scode, sxh = r_sess.text.strip().split("#", 1)
    print("scode len:", len(scode), "sxh len:", len(sxh))

    captcha = download_captcha_and_ocr(s)
    encoded = make_encoded(USERNAME, PASSWORD, scode, sxh)
    encoded_q = urllib.parse.quote_plus(encoded)
    body = f"userAccount={USERNAME}&userPassword=&RANDOMCODE={urllib.parse.quote_plus(captcha)}&encoded={encoded_q}"

    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "Origin": BASE,
        "Referer": LOGIN_PAGE,
        "User-Agent": COMMON_HEADERS["User-Agent"],
    }

    print("Step3: 发送登录请求...")
    r_post = s.post(LOGIN_POST, data=body.encode("utf-8"), headers=headers, allow_redirects=False, timeout=20)
    return s, r_post

# ---------------- 自动判断当前周 ----------------
def get_current_week():
    open_day = datetime(2025, 9, 15, tzinfo=timezone(timedelta(hours=8)))  # 开学日
    now = datetime.now(timezone(timedelta(hours=8)))
    days_diff = (now - open_day).days
    if days_diff < 0:
        return 1
    return (days_diff // 7) + 1

# ---------------- EXPORT XLS ----------------
def export_course_xls(session, login_resp):
    if 300 <= login_resp.status_code < 400 and login_resp.headers.get("Location"):
        loc = login_resp.headers["Location"]
        if loc.startswith("/"):
            loc = BASE.rstrip("/") + loc
        print(f"✅ 登录成功！访问重定向地址以激活登录态: {loc}")
        session.get(loc, headers=COMMON_HEADERS, timeout=15)

        week_number = get_current_week()
        print(f"📅 自动识别当前为第 {week_number} 周")

        params = {
            "xnxq01id": "2025-2026-1",
            "zc": str(week_number),
            "kbjcmsid": "C26030BDC5F8456CBE75B8779AED2F8A",
            "wkbkc": "1",
        }

        export_headers = {
            "Referer": f"{BASE}/jsxsd/xskb/xskb_list.do",
            "Origin": BASE,
            "User-Agent": COMMON_HEADERS["User-Agent"],
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        }

        print(f"📤 正在导出第 {week_number} 周课程表...")
        r_export = session.get(
            COURSE_EXPORT_URL,
            headers=export_headers,
            params=params,
            timeout=20
        )

        out_dir = Path("extracted_courses")
        out_dir.mkdir(exist_ok=True)

        save_path = out_dir / f"courses_week_{week_number:02}.xls"
        with open(save_path, "wb") as f:
            f.write(r_export.content)

        content_bytes = r_export.content
        if b"loginForm" in content_bytes or "请输入账号".encode("utf-8") in content_bytes:
            print("❌ 导出失败: ⚠️ 登录态失效，返回的是登录页 HTML")
            return None
        print(f"✅ 导出成功: {save_path}")
        return save_path
    else:
        print("❌ 登录未成功，请检查 debug_loginpage.html")
        return None

#-----------------XLS->XLSX--------------------
def convert_xls_to_xlsx_clean(xls_path):
    xlsx_path = xls_path.with_suffix(".xlsx")

    # 读取 xls（仅读值，不读样式）
    book = xlrd.open_workbook(xls_path, formatting_info=False)
    sheet = book.sheet_by_index(0)

    # 创建全新的 XLSX —— 无样式污染
    wb = Workbook()
    ws = wb.active

    # 设置默认列宽，使移动端可见
    for col in range(1, sheet.ncols + 1):
        column_letter = chr(64 + col)
        ws.column_dimensions[column_letter].width = 25

    # 写入内容并设置统一样式
    for r in range(sheet.nrows):
        row_values = sheet.row_values(r)
        ws.append(row_values)

        for c in range(1, len(row_values) + 1):
            cell = ws.cell(row=r+1, column=c)
            cell.font = Font(color="000000")             # 强制黑色字体
            cell.alignment = Alignment(
                wrap_text=True,                          # 自动换行
                vertical="top",
                horizontal="left"
            )

    wb.save(xlsx_path)
    return xlsx_path

#----------------清洗单元格内容：去除前置换行符------------------
def clean_xlsx_content(path):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # 自动换行、固定行高
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                cell.alignment = Alignment(
                    wrap_text=True,  # 自动换行
                    vertical="top"  # 顶部对齐
                )

    # 设置每行行高（移动端才会显示多行）
    for row in ws.iter_rows():
        row_index = row[0].row
        ws.row_dimensions[row_index].height = 110

    # 可选：列宽固定，让课程不被压扁
    for col in range(2, 9):  # 星期一到星期日
        ws.column_dimensions[get_column_letter(col)].width = 22

    wb.save(xlsx_path)
    return xlsx_path



# ---------------- MAIN ----------------
if __name__ == "__main__":
    session, login_resp = login_via_raw_body()
    xls_path = export_course_xls(session, login_resp)

    if not xls_path:
        raise SystemExit("❌ 导出失败，没有导出文件")

    xlsx_path = convert_xls_to_xlsx_clean(xls_path)
    print("转换后的文件：", xlsx_path)
    clean_xlsx_content(xlsx_path)